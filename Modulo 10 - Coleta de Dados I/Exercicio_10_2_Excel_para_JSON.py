# 2. Excel para JSON
import json
from openpyxl import load_workbook

#abrindo o arquivo excel 
planilhas = load_workbook(filename='./credito.xlsx')
planilha_ativa = planilhas.active

# Indice de cada Coluna 
cabecalho = ['escolaridade','tipo_cartao']
indice_escolaridade = None
indice_tipo_cartao = None

#cabecalho das colunas a serem extraidas 
cabecalho = next(planilha_ativa.values)
indice_escolaridade = cabecalho.index('escolaridade')
indice_tipo_cartao = cabecalho.index('tipo_cartao')

dados_escolaridade = []
dados_tipo_cartao =  []

#dados_escolaridade = [linha[indice_escolaridade] for linha in planilha_ativa if linha[indice_escolaridade] != 'escolaridade' ]

for linha in planilha_ativa.iter_rows(min_row=2, values_only=True):
    if linha[indice_escolaridade]:
        dados_escolaridade.append(linha[indice_escolaridade])
    if linha[indice_tipo_cartao]:
        dados_tipo_cartao.append(linha[indice_tipo_cartao])


dados_tipo_cartao = set(dados_tipo_cartao)
dados_escolaridade = set(dados_escolaridade)
#print(dados_escolaridade)
#print(dados_tipo_cartao)

creditos = { 'Escolaridade':
    list(dados_escolaridade),
    'Tipo_cartao':list(dados_tipo_cartao)
}

#print(creditos)

creditos_json = json.dumps(creditos,indent=4)
print(creditos_json)