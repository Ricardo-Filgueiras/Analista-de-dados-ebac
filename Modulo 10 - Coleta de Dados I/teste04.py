#testando uma Hipotese

import csv

from openpyxl import load_workbook


planilhas = load_workbook('./credito.xlsx')
planilha_01 = planilhas.active


#for linha in planilha_01.iter_rows(min_row=2):
#    if linha 