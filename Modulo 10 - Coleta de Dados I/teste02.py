import csv
import pandas as pd
from openpyxl import load_workbook


planilhas = load_workbook(filename='./credito.xlsx') #abrindo o arquivo excel
planilha_ativa = planilhas.active

# indice de cada coluna
cabecalho = ['id','sexo','idade','default','estado_civil']
indice_id = None
indice_sexo = None
indice_idade = None
indice_default = None
indice_estado_civil = None

#cabecalho das colunas a serem extraidas 

cabecalho = next(planilha_ativa.values)
indice_id = cabecalho.index('id')
indice_sexo = cabecalho.index('sexo')
indice_idade = cabecalho.index('idade')
indice_default = cabecalho.index('default')
indice_estado_civil = cabecalho.index('estado_civil')

#laco de repeticao ou loop 

dados = []
dados.append(['id','sexo','idade'])
for linha in planilha_ativa.iter_rows(min_row=2):
    if linha[indice_default].value == 1 and linha[indice_estado_civil].value == 'solteiro':
        dados.append([linha[indice_id].value,linha[indice_sexo].value,linha[indice_idade].value]) 

with open('result02.csv', mode='w', newline='') as arquivo_saida:
    escritor = csv.writer(arquivo_saida, delimiter=';')
    escritor.writerows(dados)

print(dados)
